import tarantool
import utils as utils
from openpyxl import Workbook

#--------------------------------connection management--------------------------------

def connect(db_host, db_port):
	global connection, users, tests, questions, answers, tests_results, completed_tests
	connection = tarantool.connect(db_host, db_port)
	users = connection.space('users')
	tests = connection.space('tests')
	questions = connection.space('questions')
	answers = connection.space('answers')
	tests_results = connection.space('tests_results')
	completed_tests = connection.space('completed_tests')

def disconnect():
	global connection
	connection.close()



#----------------------------------------data----------------------------------------

def update_row(table, row):
	table.update(row[0], utils.build_all_update(row))

def init_user_info(user):
	global users
	db_response = users.select(user.id, index = 0)
	#print('user info: ', db_response)
	if not db_response:
		db_response = users.insert((user.id, 0))
		#print('added info: ', db_response)

	db_info = db_response.data[0]
	user.permissions = db_info[1]
	user.state_params = {}
	#print('user info: ', user.info)

def get_admins_list():
	global users
	return users.select(2, index = 1)

class Answer:
	def __init__(self, data):
		self.data = data

	def update_info(self):
		global answers
		update_row(answers, self.data)

	def load_info(self):
		global answers
		self.data = answers.select(self.get_id())[0]

	def get_id(self):
		return self.data[0]

	def get_text(self):
		return self.data[2]

	def set_text(self, value):
		self.data[2] = value

	def get_image(self):
		result = self.data[3]
		if result == '':
			return None
		return result

	def set_image(self, value):
		if value == None:
			value = ''
		self.data[3] = value

	def get_value(self):
		return self.data[4]

	def set_value(self, value):
		self.data[4] = value

	def remove(self):
		remove_answer(self.get_id())

	def get_formatted_text(self):
		if self.get_image(): 
			return self.get_text() + ' ' + utils.get_reference_to_file(self.get_image()) + ' '
		else:
			return self.get_text()
		
class Quesion:
	def __init__(self, data):
		self.data = data
		self.load_info(False)

	def load_info(self, with_data = True):
		global questions, answers

		if with_data:
			self.data = questions.select(self.get_id())[0]

		self.answers = []
		for row in answers.select(self.get_id(), index = 1):
			self.answers.append(Answer(row))

		self.answers.sort(key = (lambda answer : answer.get_id()))

	def get_answers(self):
		return self.answers

	def update_info(self):
		global questions
		update_row(questions, self.data)

	def get_id(self):
		return self.data[0]

	def get_text(self):
		return self.data[2]

	def set_text(self, value):
		self.data[2] = value

	def get_image(self):
		result = self.data[3]
		if result == '':
			return None
		return result

	def set_image(self, value):
		if value == None:
			value = ''
		self.data[3] = value

	def remove(self):
		remove_question(self.get_id())

	def has_image_answers(self):
		for answer in self.get_answers():
			if answer.get_image():
				return True
		return False

	def get_formatted_text(self):
		if self.get_image(): 
			return self.get_text() + ' ' + utils.get_reference_to_file(self.get_image()) + ' '
		else:
			return self.get_text()

class TestResult:
	def __init__(self, data):
		self.data = data

	def update_info(self):
		global tests_results
		update_row(tests_results, self.data)

	def load_info(self):
		global tests_results
		self.data = tests_results.select(self.get_id())[0]

	def get_id(self):
		return self.data[0]

	def get_text(self):
		return self.data[2]

	def set_text(self, value):
		self.data[2] = value
	
	def get_image(self):
		result = self.data[3]
		if result == '':
			return None
		return result

	def set_image(self, value):
		if value == None:
			value = ''
		self.data[3] = value


	def get_value(self):
		return self.data[4]

	def set_value(self, value):
		self.data[4] = value
	
	def remove(self):
		remove_test_result(self.get_id())

	def get_formatted_text(self):
		if self.get_image(): 
			return self.get_text() + ' ' + utils.get_reference_to_file(self.get_image()) + ' '
		else:
			return self.get_text()

class Test:
	def __init__(self, data):
		self.data = data
		self.load_info(False)

	def load_info(self, with_data = True):
		global tests, questions, tests_results

		if with_data:
			self.data = tests.select(self.get_id())[0]

		self.questions = []
		questions_result = questions.select(self.get_id(), index = 1)
		for row in questions_result:
			self.questions.append(Quesion(row))

		self.questions.sort(key = (lambda question : question.get_id()))
		
		
		self.tests_results = []
		for row in tests_results.select(self.get_id(), index = 1):
			self.tests_results.append(TestResult(row))

		self.tests_results.sort(key = (lambda test_result : test_result.get_id()))

	def update_info(self):
		global tests
		update_row(tests, self.data)

	def get_id(self):
		return self.data[0]

	def get_name(self):
		return self.data[1]

	def get_handle_module(self):
		return self.data[2]

	def set_handler_module(self, name):
		global tests
		self.data[2] = name
	
	def get_publish_status(self):
		return self.data[3]

	def set_publish_status(self, status):
		self.data[3] = status

	def get_questions(self):
		return self.questions

	def get_results(self):
		return self.tests_results

	def remove(self):
		self.set_publish_status(2)
		self.update_info()

	def can_complete(self, user):
		if user.permissions > 0:
			return True
		return self.get_publish_status()

class TestsList:
	def __init__(self):
		self.load_tests()
	
	def get_all_tests(self, include_deleted = False): 
		result = []
		for test in self.tests:
			if include_deleted or test.get_publish_status() != 2:
				result.append(test)

		return result

	def get_published_tests(self):
		result = []
		for test in self.tests:
			if test.get_publish_status() == 1:
				result.append(test)

		return result

	def load_tests(self):
		global tests

		self.tests = []
		tests_result = tests.select()
		for cortage in tests_result:
			self.tests.append(Test(cortage))

		self.tests.sort(key = (lambda test : test.get_id()))
		#self.set_numbers()
	
def load_tests_list_clone():
	tests_list = TestsList()
	tests_list.load_tests()

	return tests_list
	
def update_tests():
	global tests_list
	tests_list = load_tests_list_clone()

def add_test(name):
	connection.call('add_test', name)

def add_question(test_id, text, image):
	if not image:
		image = ''
	connection.call('add_question', test_id, text, image)

def add_test_result(test_id, text, image, value):
	if not image:
		image = ''
	connection.call('add_test_result', test_id, text, image, value)

def add_answer(question_id, text, image, value):
	if not image:
		image = ''
	connection.call('add_answer', question_id, text, image, value)

def remove_question(id):
	global questions
	questions.delete(id)

def remove_test_result(id):
	global tests_results
	tests_results.delete(id)

def remove_answer(id):
	global answers
	answers.delete(id)

def update_user_by_id(user_cortage):
	global users
	users.replace(user_cortage)

def add_user_test_complete(user_id, test_id):
	global completed_tests
	completed_tests.upsert((user_id, test_id, 1), [('+', 2, 1)])

class TestStatisticsObject:
	
	def __init__(self, sheet):
		self.sheet = sheet
		self.sheet['A1'] = 'UIN'
		self.sheet['B1'] = 'Кол-во'
		self.sheet.column_dimensions['A'].width = 25
		self.sheet.column_dimensions['B'].width = 25
		self.current_row = 3
		self.all_count = 0
		self.rows_count = 0

	def add(self, user_id, count):
		self.sheet['A%d'%(self.current_row)] = user_id
		self.sheet['B%d'%(self.current_row)] = count
		self.current_row += 1
		self.all_count += count
		self.rows_count += 1

	def end(self):
		self.sheet['A2'] = self.rows_count
		self.sheet['B2'] = self.all_count


def get_statistics():
	global tests, completed_tests
	wb = Workbook()
	wb.remove(wb.active)

	#create worksheets
	tests_statistics = dict()
	tests_list = load_tests_list_clone()
	sheet_counter = 0
	for test in tests_list.get_all_tests(True):
		tests_statistics[test.get_id()] = TestStatisticsObject(wb.create_sheet(test.get_name(), sheet_counter))
		sheet_counter += 1

	common_statistics = TestStatisticsObject(wb.create_sheet('Все тесты', sheet_counter))
	sheet_counter += 1



	common_statistics_dictionary = dict()
	for row in completed_tests.select():
		user_id = row[0]
		test_id = row[1]
		count = row[2]
		
		test_statistic_object = tests_statistics.get(test_id, None)
		if test_statistic_object:
			test_statistic_object.add(user_id, count)

		#common statistics
		if user_id in common_statistics_dictionary:
			common_statistics_dictionary[user_id] += count
		else:
			common_statistics_dictionary[user_id] = count

	for user_id, count in common_statistics_dictionary.items():
		common_statistics.add(user_id, count)

	
	

	for tso in tests_statistics.values():
		tso.end()

	common_statistics.end()
	
	return wb

class FastStatisticsObject:

	def __init__(self, test_name):
		self.test_name = test_name
		self.count = 0
		self.users = 0
	
	def add(self, user_id, count):
		self.count += count
		self.users += 1

def get_fast_statistics():
	global tests, completed_tests

	#create worksheets
	tests_statistics = dict()
	tests_list = load_tests_list_clone()
	for test in tests_list.get_all_tests(True):
		tests_statistics[test.get_id()] = FastStatisticsObject(test.get_name())

	common_statistics = FastStatisticsObject('Всего')
	common_statistics_dictionary = dict()

	for row in completed_tests.select():
		user_id = row[0]
		test_id = row[1]
		count = row[2]

		test_statistic_object = tests_statistics.get(test_id, None)
		if test_statistic_object:
			test_statistic_object.add(user_id, count)

		#common statistics
		if user_id in common_statistics_dictionary:
			common_statistics_dictionary[user_id] += count
		else:
			common_statistics_dictionary[user_id] = count
	
	for user_id, count in common_statistics_dictionary.items():
		common_statistics.add(user_id, count)


	return tests_statistics, common_statistics

def get_users():
	global users
	return users.select()